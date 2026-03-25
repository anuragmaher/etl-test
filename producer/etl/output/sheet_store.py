"""Store sheet data as CSV + SQLite for structured querying."""

import csv
import io
import logging
import os
import re
import sqlite3

from openpyxl import load_workbook

logger = logging.getLogger(__name__)


def sanitize_table_name(name: str) -> str:
    """Create a safe SQLite table name."""
    name = re.sub(r'[^\w]', '_', name).strip('_')
    return name[:100] or "sheet"


def sanitize_column_name(name: str) -> str:
    """Create a safe SQLite column name."""
    name = re.sub(r'[^\w]', '_', str(name)).strip('_')
    return name[:100] or "col"


def store_sheet(xlsx_bytes: bytes, doc_id: str, title: str, data_dir: str = "./data") -> dict:
    """Store sheet data as CSV files and in SQLite. Returns schema info for Pinecone metadata."""
    os.makedirs(data_dir, exist_ok=True)
    db_path = os.path.join(data_dir, "sheets.db")

    wb = load_workbook(io.BytesIO(xlsx_bytes), read_only=True, data_only=True)
    schema_info = {"title": title, "doc_id": doc_id, "sheets": []}

    conn = sqlite3.connect(db_path)

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        rows = [r for r in rows if any(cell is not None for cell in r)]

        if not rows or len(rows) < 2:
            continue

        # First row is headers
        headers = [str(cell) if cell is not None else f"col_{i}" for i, cell in enumerate(rows[0])]
        data_rows = rows[1:]

        # Create a unique table name
        table_name = sanitize_table_name(f"{sanitize_table_name(doc_id[:20])}_{sanitize_table_name(sheet_name)}")

        # Save as CSV
        csv_path = os.path.join(data_dir, f"{table_name}.csv")
        with open(csv_path, "w", newline="") as f:
            writer = csv.writer(f)
            writer.writerow(headers)
            for row in data_rows:
                writer.writerow([str(cell) if cell is not None else "" for cell in row])

        # Create SQLite table
        safe_columns = [sanitize_column_name(h) for h in headers]
        # Drop existing table
        conn.execute(f'DROP TABLE IF EXISTS "{table_name}"')

        # Detect column types from data
        col_defs = []
        for i, col in enumerate(safe_columns):
            # Check if all values in this column are numeric
            is_numeric = True
            for row in data_rows[:50]:  # Check first 50 rows
                val = row[i] if i < len(row) else None
                if val is not None and val != "":
                    try:
                        float(str(val).replace(",", ""))
                    except (ValueError, TypeError):
                        is_numeric = False
                        break
            col_type = "REAL" if is_numeric else "TEXT"
            col_defs.append(f'"{col}" {col_type}')

        create_sql = f'CREATE TABLE "{table_name}" ({", ".join(col_defs)})'
        conn.execute(create_sql)

        # Insert data
        placeholders = ", ".join(["?"] * len(safe_columns))
        insert_sql = f'INSERT INTO "{table_name}" VALUES ({placeholders})'
        for row in data_rows:
            values = []
            for i, cell in enumerate(row):
                if i >= len(safe_columns):
                    break
                if cell is None:
                    values.append(None)
                else:
                    values.append(str(cell))
            # Pad with None if row is shorter than headers
            while len(values) < len(safe_columns):
                values.append(None)
            conn.execute(insert_sql, values)

        conn.commit()

        # Build schema info
        sample_rows = data_rows[:3]
        sample_text = []
        for row in sample_rows:
            row_dict = {h: str(row[i]) if i < len(row) and row[i] is not None else ""
                        for i, h in enumerate(headers)}
            sample_text.append(str(row_dict))

        sheet_info = {
            "sheet_name": sheet_name,
            "table_name": table_name,
            "columns": headers,
            "row_count": len(data_rows),
            "sample_rows": sample_text,
        }
        schema_info["sheets"].append(sheet_info)

        logger.info("Stored sheet '%s.%s' as table '%s' (%d rows, %d columns)",
                     title, sheet_name, table_name, len(data_rows), len(headers))

    conn.close()
    wb.close()
    return schema_info


def query_sheet(sql: str, data_dir: str = "./data") -> dict:
    """Execute a SQL query against the sheets database. Returns {columns, rows, error}."""
    db_path = os.path.join(data_dir, "sheets.db")
    if not os.path.exists(db_path):
        return {"columns": [], "rows": [], "error": "No sheets database found."}

    try:
        conn = sqlite3.connect(db_path)
        cursor = conn.execute(sql)
        columns = [desc[0] for desc in cursor.description] if cursor.description else []
        rows = cursor.fetchall()
        conn.close()
        return {"columns": columns, "rows": rows, "error": None}
    except Exception as e:
        return {"columns": [], "rows": [], "error": str(e)}


def list_tables(data_dir: str = "./data") -> list:
    """List all tables and their schemas in the sheets database."""
    db_path = os.path.join(data_dir, "sheets.db")
    if not os.path.exists(db_path):
        return []

    conn = sqlite3.connect(db_path)
    cursor = conn.execute("SELECT name FROM sqlite_master WHERE type='table'")
    tables = []
    for (table_name,) in cursor.fetchall():
        col_cursor = conn.execute(f'PRAGMA table_info("{table_name}")')
        columns = [(row[1], row[2]) for row in col_cursor.fetchall()]
        count_cursor = conn.execute(f'SELECT COUNT(*) FROM "{table_name}"')
        row_count = count_cursor.fetchone()[0]
        tables.append({
            "table_name": table_name,
            "columns": columns,
            "row_count": row_count,
        })
    conn.close()
    return tables
