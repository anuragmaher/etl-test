"""Convert an Excel/Google Sheet (.xlsx bytes) to Markdown."""

import io

from openpyxl import load_workbook


def convert(xlsx_bytes: bytes) -> str:
    """Convert .xlsx file bytes to a Markdown string.

    Each sheet becomes a section with a heading and a markdown table.
    """
    wb = load_workbook(io.BytesIO(xlsx_bytes), read_only=True, data_only=True)
    sections = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))

        if not rows:
            continue

        # Filter out completely empty rows
        rows = [r for r in rows if any(cell is not None for cell in r)]
        if not rows:
            continue

        # Build markdown table
        md_rows = []
        for i, row in enumerate(rows):
            cells = [str(cell) if cell is not None else "" for cell in row]
            md_rows.append("| " + " | ".join(cells) + " |")
            if i == 0:
                md_rows.append("| " + " | ".join(["---"] * len(cells)) + " |")

        section = f"## {sheet_name}\n\n" + "\n".join(md_rows)
        sections.append(section)

    wb.close()

    if not sections:
        return "\n"

    # If only one sheet, skip the heading
    if len(sections) == 1:
        # Remove the "## SheetName\n\n" prefix
        result = sections[0].split("\n\n", 1)[1] if "\n\n" in sections[0] else sections[0]
    else:
        result = "\n\n".join(sections)

    return result.strip() + "\n"
